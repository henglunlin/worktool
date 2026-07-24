# -*- coding: utf-8 -*-

import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# =========================================================
# 常數
# =========================================================

DNI_COMPARE_COL = "__DNI_COMPARE__"


# =========================================================
# 頁面設定
# =========================================================

from PIL import Image
icon = Image.open("logo.png")

st.set_page_config(
    page_title="BOM 比對工具",
    page_icon=icon,
    layout="wide",
)

import streamlit as st

# 1. 關鍵修改：加入 vertical_alignment="center" 讓兩欄內容垂直置中對齊
col1, col2 = st.columns([0.5, 10], vertical_alignment="center")

with col1:
    # 2. 移除 width=50，改讓欄位比例 [1, 10] 或使用 use_container_width=True 自動適應欄寬
    st.image("logo.png", use_container_width=True)

with col2:
    # 3. 關鍵修改：使用 markdown 替代 title，並透過內聯 CSS 消除自帶的上下邊距 (margin: 0)
    st.markdown(
        "<h1 style='margin: 0; padding: 0;'>BOM 比對工具-告訴我你會買日月光</h1>",help="Alan 祝您鑽大錢", 
        unsafe_allow_html=True
    )



# =========================================================
# 基本工具函數
# =========================================================

def show_friendly_error(title, message, detail=None):
    st.error(title)
    st.warning(message)

    if detail is not None:
        with st.expander("查看技術細節"):
            st.code(str(detail))

    st.stop()


def clean_col_name(col):
    if pd.isna(col):
        return ""

    col = str(col).strip()
    col = col.replace("\n", " ")
    col = col.replace("\r", " ")
    col = " ".join(col.split())

    return col


def clean_value(value, question_as_blank=True):
    if pd.isna(value):
        return ""

    value = str(value).strip()
    value = " ".join(value.split())

    if question_as_blank:
        blank_tokens = {
            "",
            "?",
            "？",
            "nan",
            "none",
            "null",
            "n/a",
            "na",
            "#n/a",
            "#na",
        }

        if value.lower() in blank_tokens:
            return ""

    return value


def normalize_compare_value(value, case_sensitive=False):
    value = clean_value(value, question_as_blank=True)
    value = " ".join(value.split())

    if not case_sensitive:
        value = value.upper()

    return value


def is_dni(value):
    value = clean_value(value, question_as_blank=True).upper()
    return "DNI" in value


def contains_chinese(text):
    text = clean_value(text, question_as_blank=True)
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def excel_font_by_value(value, bold=False, color="000000"):
    value = clean_value(value, question_as_blank=True)

    if contains_chinese(value):
        font_name = "微軟正黑體"
    else:
        font_name = "Calibri"

    return Font(
        name=font_name,
        bold=bold,
        color=color,
    )


def natural_sort_key(text):
    text = clean_value(text, question_as_blank=True)
    parts = re.split(r"(\d+)", text)

    sort_key = []

    for part in parts:
        if part.isdigit():
            sort_key.append((1, int(part)))
        else:
            sort_key.append((0, part.upper()))

    return sort_key


def guess_column(columns, candidates):
    upper_map = {str(c).upper(): c for c in columns}

    for cand in candidates:
        if cand.upper() in upper_map:
            return upper_map[cand.upper()]

    for col in columns:
        col_upper = str(col).upper()

        for cand in candidates:
            cand_upper = cand.upper()

            if cand_upper in col_upper or col_upper in cand_upper:
                return col

    return None


def is_possible_dni_column(col_name):
    col_upper = str(col_name).upper()

    return (
        "IGNORE" in col_upper
        or "DNI" in col_upper
        or "MOUNT" in col_upper
        or "BOM_IGNORE" in col_upper
        or "BOM IGNORE" in col_upper
    )


# =========================================================
# 檔案讀取工具：CSV / XLS / XLSX / XLSM 
# =========================================================

def get_uploaded_file_ext(uploaded_file):
    file_name = uploaded_file.name.lower().strip()

    if "." not in file_name:
        return ""

    return file_name.rsplit(".", 1)[-1]


def get_excel_engine_by_ext(ext):
    if ext in ["xlsx", "xlsm"]:
        return "openpyxl"

    if ext == "xls":
        return "xlrd"

    raise ValueError("不支援的 Excel 副檔名：{}".format(ext))


def read_csv_with_fallback(uploaded_file, header=None):
    encodings = [
        "utf-8-sig",
        "utf-8",
        "cp950",
        "big5",
        "latin1",
    ]

    last_error = None

    for encoding in encodings:
        try:
            uploaded_file.seek(0)

            return pd.read_csv(
                uploaded_file,
                header=header,
                dtype=str,
                encoding=encoding,
                sep=None,
                engine="python",
            )

        except Exception as e:
            last_error = e

    raise last_error


def read_input_sheets(uploaded_file):
    """
    CSV 沒有 Sheet，固定回傳 ['CSV']。
    Excel 回傳實際 Sheet 名稱。
    """
    ext = get_uploaded_file_ext(uploaded_file)

    if ext == "csv":
        return ["CSV"]

    if ext in ["xls", "xlsx", "xlsm"]:
        uploaded_file.seek(0)
        engine = get_excel_engine_by_ext(ext)

        xls = pd.ExcelFile(
            uploaded_file,
            engine=engine,
        )

        return xls.sheet_names

    raise ValueError(
        "不支援的檔案格式：.{}\n\n目前支援：csv、xls、xlsx、xlsm".format(ext)
    )


def read_table_raw(uploaded_file, sheet_name=None):
    """
    讀取原始資料，不指定 Header。
    """
    ext = get_uploaded_file_ext(uploaded_file)

    if ext == "csv":
        return read_csv_with_fallback(
            uploaded_file,
            header=None,
        )

    if ext in ["xls", "xlsx", "xlsm"]:
        uploaded_file.seek(0)
        engine = get_excel_engine_by_ext(ext)

        return pd.read_excel(
            uploaded_file,
            sheet_name=sheet_name,
            header=None,
            engine=engine,
            dtype=str,
        )

    raise ValueError(
        "不支援的檔案格式：.{}\n\n目前支援：csv、xls、xlsx、xlsm".format(ext)
    )


def detect_header_row(raw_df):
    keywords = [
        "REF DES",
        "REFERENCE DESIGNATOR",
        "DESIGNATOR",
        "PN",
        "P/N",
        "PART",
        "PART NUMBER",
        "MFR",
        "MFG",
        "MANUFACTURER",
        "QTY",
        "BOM_IGNORE",
        "BOM IGNORE",
        "DNI",
    ]

    best_row = 0
    best_score = -1
    max_scan_rows = min(len(raw_df), 80)

    for idx in range(max_scan_rows):
        row_values = [
            clean_col_name(v).upper()
            for v in raw_df.iloc[idx].tolist()
            if clean_value(v, question_as_blank=True) != ""
        ]

        score = 0

        for cell in row_values:
            for keyword in keywords:
                if keyword in cell:
                    score += 1

        non_empty_count = len(row_values)

        if non_empty_count >= 3:
            score += min(non_empty_count, 10) * 0.1

        if score > best_score:
            best_score = score
            best_row = idx

    return best_row


def load_bom_dataframe(uploaded_file, sheet_name, header_row):
    """
    CSV：
    - sheet_name 會忽略
    - header_row 一樣有效

    Excel：
    - 依指定 sheet_name / header_row 讀取
    """
    ext = get_uploaded_file_ext(uploaded_file)

    if ext == "csv":
        df = read_csv_with_fallback(
            uploaded_file,
            header=header_row,
        )

    elif ext in ["xls", "xlsx", "xlsm"]:
        uploaded_file.seek(0)
        engine = get_excel_engine_by_ext(ext)

        df = pd.read_excel(
            uploaded_file,
            sheet_name=sheet_name,
            header=header_row,
            engine=engine,
            dtype=str,
        )

    else:
        raise ValueError(
            "不支援的檔案格式：.{}\n\n目前支援：csv、xls、xlsx、xlsm".format(ext)
        )

    df.columns = [clean_col_name(c) for c in df.columns]

    valid_cols = []

    for c in df.columns:
        c_text = clean_value(c, question_as_blank=True)

        if c_text != "" and not c_text.upper().startswith("UNNAMED"):
            valid_cols.append(c)

    df = df.loc[:, valid_cols]
    df = df.dropna(how="all")

    for col in df.columns:
        df[col] = df[col].apply(lambda x: clean_value(x, question_as_blank=True))

    return df


# =========================================================
# BOM 前處理
# =========================================================

def explode_key_column(df, key_col, separator=","):
    df = df.copy()

    df[key_col] = df[key_col].fillna("").astype(str).str.split(separator)
    df = df.explode(key_col).copy()
    df[key_col] = df[key_col].apply(lambda x: clean_value(x, question_as_blank=True))

    df = df[
        (df[key_col] != "")
        & (df[key_col].str.lower() != "nan")
        & (df[key_col] != "?")
        & (df[key_col] != "？")
    ].copy()

    return df


def deduplicate_by_key(df, key_col, compare_cols):
    df = df.copy()

    agg_dict = {}

    for col in compare_cols:

        def merge_values(series):
            values = []

            for v in series.tolist():
                v = clean_value(v, question_as_blank=True)

                if v != "":
                    values.append(v)

            unique_values = sorted(set(values))
            return " | ".join(unique_values)

        agg_dict[col] = merge_values

    out = df.groupby(key_col, dropna=False, as_index=False).agg(agg_dict)

    return out


def prepare_bom(
    df,
    key_col,
    compare_cols,
    explode_key=True,
    remove_blank_key=True,
    extra_cols=None,
):
    if extra_cols is None:
        extra_cols = []

    df = df.copy()

    selected_cols = list(dict.fromkeys([key_col] + compare_cols + extra_cols))
    selected_cols = [col for col in selected_cols if col in df.columns]

    df = df[selected_cols].copy()

    for col in selected_cols:
        df[col] = df[col].apply(lambda x: clean_value(x, question_as_blank=True))

    if remove_blank_key:
        df = df[
            (df[key_col] != "")
            & (df[key_col].str.lower() != "nan")
            & (df[key_col] != "?")
            & (df[key_col] != "？")
        ].copy()

    if explode_key:
        df = explode_key_column(df, key_col)

    group_cols = list(dict.fromkeys(compare_cols + extra_cols))
    group_cols = [col for col in group_cols if col in df.columns]

    df = deduplicate_by_key(df, key_col, group_cols)

    try:
        df = df.sort_values(
            by=key_col,
            key=lambda s: s.map(natural_sort_key),
        ).reset_index(drop=True)
    except Exception:
        df = df.sort_values(by=key_col).reset_index(drop=True)

    return df


# =========================================================
# BOM 比對邏輯
# =========================================================

def compare_bom(
    old_df,
    new_df,
    key_col,
    compare_cols,
    case_sensitive=False,
    dni_compare=False,
    bom_ignore_col=None,
):
    old_map = old_df.set_index(key_col).to_dict(orient="index")
    new_map = new_df.set_index(key_col).to_dict(orient="index")

    all_keys = sorted(
        set(old_map.keys()) | set(new_map.keys()),
        key=natural_sort_key,
    )

    result_rows = []
    detail_rows = []

    for key in all_keys:
        in_old = key in old_map
        in_new = key in new_map

        old_info = old_map.get(key, {})
        new_info = new_map.get(key, {})

        old_is_dni = False
        new_is_dni = False

        if dni_compare and bom_ignore_col:
            if in_old:
                old_is_dni = is_dni(old_info.get(bom_ignore_col, ""))
            if in_new:
                new_is_dni = is_dni(new_info.get(bom_ignore_col, ""))

        changed_cols = []

        if in_old and in_new:
            for col in compare_cols:
                old_val = old_info.get(col, "")
                new_val = new_info.get(col, "")

                old_norm = normalize_compare_value(
                    old_val,
                    case_sensitive=case_sensitive,
                )
                new_norm = normalize_compare_value(
                    new_val,
                    case_sensitive=case_sensitive,
                )

                if old_norm != new_norm:
                    changed_cols.append(col)

        is_same = len(changed_cols) == 0

        if not dni_compare:
            if in_old and in_new:
                if is_same:
                    status = "Unchanged"
                else:
                    status = "Changed"
            elif in_old and not in_new:
                status = "Del"
            elif not in_old and in_new:
                status = "New Add"
            else:
                status = "Unchanged"

        else:
            if in_old and in_new:
                if old_is_dni and not new_is_dni:
                    status = "New Add"
                elif not old_is_dni and new_is_dni:
                    status = "Del"
                elif old_is_dni and new_is_dni:
                    status = "Unchanged"
                else:
                    if is_same:
                        status = "Unchanged"
                    else:
                        status = "Changed"

            elif in_old and not in_new:
                if old_is_dni:
                    status = "Unchanged"
                else:
                    status = "Del"

            elif not in_old and in_new:
                if new_is_dni:
                    status = "Unchanged"
                else:
                    status = "New Add"

            else:
                status = "Unchanged"

        row = {
            key_col: key,
            "狀態": status,
            "舊版是否DNI": "Y" if old_is_dni else "",
            "新版是否DNI": "Y" if new_is_dni else "",
        }

        changed_col_names = []

        for col in compare_cols:
            old_val = clean_value(old_info.get(col, ""), question_as_blank=True) if in_old else ""
            new_val = clean_value(new_info.get(col, ""), question_as_blank=True) if in_new else ""

            row["舊版_" + col] = old_val
            row["新版_" + col] = new_val

            diff_status = ""

            if status == "Del":
                diff_status = "Del"
            elif status == "New Add":
                diff_status = "New Add"
            elif status == "Changed" and col in changed_cols:
                diff_status = "Diff"

            row[col + "_差異"] = diff_status

            if diff_status != "":
                changed_col_names.append(col)

                detail_rows.append(
                    {
                        key_col: key,
                        "狀態": status,
                        "差異欄位": col,
                        "舊版值": old_val,
                        "新版值": new_val,
                        "差異結果": diff_status,
                        "舊版是否DNI": "Y" if old_is_dni else "",
                        "新版是否DNI": "Y" if new_is_dni else "",
                    }
                )

        row["差異欄位清單"] = ", ".join(changed_col_names)
        result_rows.append(row)

    result_df = pd.DataFrame(result_rows)

    detail_columns = [
        key_col,
        "狀態",
        "差異欄位",
        "舊版值",
        "新版值",
        "差異結果",
        "舊版是否DNI",
        "新版是否DNI",
    ]

    detail_df = pd.DataFrame(detail_rows, columns=detail_columns)

    return result_df, detail_df


def make_change_list(result_df):
    if result_df is None or result_df.empty:
        return pd.DataFrame()

    change_list_df = result_df[
        result_df["狀態"].isin(["Changed", "New Add", "Del"])
    ].copy()

    return change_list_df


# =========================================================
# Summary 與 Excel 輸出
# =========================================================

def make_summary(result_df, old_file_name="", new_file_name="", old_sheet_name="", new_sheet_name=""):
    total = len(result_df)
    counts = result_df["狀態"].value_counts().to_dict()

    # 將本次比對使用的檔案名稱寫入輸出的 Excel Summary 工作表
    summary_data = [
        {"項目": "舊版 BOM 檔名", "數量": old_file_name},
        {"項目": "新版 BOM 檔名", "數量": new_file_name},
        {"項目": "舊版 BOM Sheet", "數量": old_sheet_name},
        {"項目": "新版 BOM Sheet", "數量": new_sheet_name},
        {"項目": "總筆數", "數量": total},
        {"項目": "Changed", "數量": counts.get("Changed", 0)},
        {"項目": "New Add", "數量": counts.get("New Add", 0)},
        {"項目": "Del", "數量": counts.get("Del", 0)},
        {"項目": "Unchanged", "數量": counts.get("Unchanged", 0)},
        {"項目": "產生時間", "數量": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]

    return pd.DataFrame(summary_data)


def clean_dataframe_for_output(df):
    df = df.copy()

    for col in df.columns:
        df[col] = df[col].apply(lambda x: clean_value(x, question_as_blank=True))

    return df


def style_excel_output(result_df, change_list_df, detail_df, summary_df):
    result_df = clean_dataframe_for_output(result_df)
    change_list_df = clean_dataframe_for_output(change_list_df)
    detail_df = clean_dataframe_for_output(detail_df)
    summary_df = clean_dataframe_for_output(summary_df)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        change_list_df.to_excel(writer, sheet_name="Change_List", index=False)
        result_df.to_excel(writer, sheet_name="Compare_Result", index=False)
        detail_df.to_excel(writer, sheet_name="Diff_Detail", index=False)

    output.seek(0)
    wb = load_workbook(output)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    changed_fill = PatternFill("solid", fgColor="FFF2CC")
    new_add_fill = PatternFill("solid", fgColor="F4B084")
    del_fill = PatternFill("solid", fgColor="FF9999")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fill_map = {
        "Unchanged": white_fill,
        "Changed": changed_fill,
        "New Add": new_add_fill,
        "Del": del_fill,
    }

    change_list_status_font_color_map = {
        "Del": "FF0000",
        "New Add": "00B050",
        "Changed": "0000FF",
    }

    for ws in wb.worksheets:
        ws.row_dimensions[1].height = 30

        status_col_idx = None

        for idx, cell in enumerate(ws[1], start=1):
            if clean_value(cell.value, question_as_blank=True) == "狀態":
                status_col_idx = idx
                break

        # Header
        for cell in ws[1]:
            if ws.title == "Change_List":
                cell.fill = white_fill
            else:
                cell.fill = header_fill

            cell.font = excel_font_by_value(
                cell.value,
                bold=True,
                color="000000",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = thin_border

        # =================================================
        # Change_List：
        # 底色白色，只改狀態欄文字顏色
        # =================================================
        if ws.title == "Change_List":
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                row_status = ""

                if status_col_idx:
                    row_status = clean_value(
                        row[status_col_idx - 1].value,
                        question_as_blank=True,
                    )

                for cell in row:
                    cell.fill = white_fill
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.border = thin_border

                    font_color = "000000"

                    if status_col_idx and cell.column == status_col_idx:
                        font_color = change_list_status_font_color_map.get(
                            row_status,
                            "000000",
                        )

                    cell.font = excel_font_by_value(
                        cell.value,
                        bold=False,
                        color=font_color,
                    )

        # =================================================
        # Compare_Result：
        # 預設白色，再依照「狀態」欄整列上色到最後欄
        #
        # New Add   => 整列橘色
        # Unchanged => 整列白色
        # Changed   => 整列黃色
        # Del       => 整列紅色
        # =================================================
        elif ws.title == "Compare_Result":
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                row_status = ""

                if status_col_idx:
                    row_status = clean_value(
                        row[status_col_idx - 1].value,
                        question_as_blank=True,
                    )

                row_fill = status_fill_map.get(row_status, white_fill)

                for cell in row:
                    cell.fill = row_fill
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.border = thin_border
                    cell.font = excel_font_by_value(
                        cell.value,
                        bold=False,
                        color="000000",
                    )

        # =================================================
        # Diff_Detail：
        # 維持依狀態整列上色
        # =================================================
        elif ws.title == "Diff_Detail":
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                row_status = ""

                if status_col_idx:
                    row_status = clean_value(
                        row[status_col_idx - 1].value,
                        question_as_blank=True,
                    )

                row_fill = status_fill_map.get(row_status, white_fill)

                for cell in row:
                    cell.fill = row_fill
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.border = thin_border
                    cell.font = excel_font_by_value(
                        cell.value,
                        bold=False,
                        color="000000",
                    )

        # =================================================
        # Summary：
        # 白底
        # =================================================
        else:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.fill = white_fill
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.border = thin_border
                    cell.font = excel_font_by_value(
                        cell.value,
                        bold=False,
                        color="000000",
                    )

        # 欄寬、自動篩選、凍結窗格
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                value = clean_value(cell.value, question_as_blank=True)

                if len(value) > max_length:
                    max_length = len(value)

            adjusted_width = min(max(max_length + 2, 12), 60)
            ws.column_dimensions[col_letter].width = adjusted_width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output


# =========================================================
# 欄位標準化與錯誤檢查
# =========================================================

def rename_for_compare(
    old_df,
    new_df,
    old_key,
    new_key,
    mapping,
    extra_keep_cols=None,
):
    if extra_keep_cols is None:
        extra_keep_cols = []

    old_df = old_df.copy()
    new_df = new_df.copy()

    compare_key = "Compare_Key"

    if old_key not in old_df.columns:
        raise ValueError(
            "舊版 Key 欄位不存在：{}\n\n舊版目前可用欄位：{}".format(
                old_key,
                list(old_df.columns),
            )
        )

    if new_key not in new_df.columns:
        raise ValueError(
            "新版 Key 欄位不存在：{}\n\n新版目前可用欄位：{}".format(
                new_key,
                list(new_df.columns),
            )
        )

    missing_old_cols = [
        old_col for old_col in mapping.keys()
        if old_col not in old_df.columns
    ]

    if missing_old_cols:
        raise ValueError(
            "以下舊版比對欄位不存在：{}\n\n舊版目前可用欄位：{}".format(
                missing_old_cols,
                list(old_df.columns),
            )
        )

    missing_new_cols = [
        new_col for new_col in mapping.values()
        if new_col not in new_df.columns
    ]

    if missing_new_cols:
        raise ValueError(
            "以下新版比對欄位不存在：{}\n\n新版目前可用欄位：{}".format(
                missing_new_cols,
                list(new_df.columns),
            )
        )

    selected_new_cols = list(mapping.values())

    duplicated_new_cols = sorted(
        {
            col for col in selected_new_cols
            if selected_new_cols.count(col) > 1
        }
    )

    if duplicated_new_cols:
        duplicate_messages = []

        for new_col in duplicated_new_cols:
            related_old_cols = [
                old_col
                for old_col, mapped_new_col in mapping.items()
                if mapped_new_col == new_col
            ]

            duplicate_messages.append(
                "新版欄位 [{}] 被以下舊版欄位重複使用：{}".format(
                    new_col,
                    ", ".join(related_old_cols),
                )
            )

        raise ValueError(
            "欄位對應錯誤：同一個新版欄位不可被多個舊版欄位重複使用。\n\n"
            + "\n".join(duplicate_messages)
            + "\n\n請回到欄位對應區，重新選擇正確的新版欄位。"
        )

    old_rename = {
        old_key: compare_key,
    }

    new_rename = {
        new_key: compare_key,
    }

    compare_cols = []

    for old_col, new_col in mapping.items():
        common_name = old_col
        old_rename[old_col] = common_name
        new_rename[new_col] = common_name
        compare_cols.append(common_name)

    old_df = old_df.rename(columns=old_rename)
    new_df = new_df.rename(columns=new_rename)

    old_needed = list(dict.fromkeys([compare_key] + compare_cols + extra_keep_cols))
    new_needed = list(dict.fromkeys([compare_key] + compare_cols + extra_keep_cols))

    missing_old_after_rename = [
        col for col in old_needed
        if col not in old_df.columns
    ]

    missing_new_after_rename = [
        col for col in new_needed
        if col not in new_df.columns
    ]

    if missing_old_after_rename:
        raise ValueError(
            "舊版欄位重新命名後缺少欄位：{}\n\n舊版目前欄位：{}".format(
                missing_old_after_rename,
                list(old_df.columns),
            )
        )

    if missing_new_after_rename:
        raise ValueError(
            "新版欄位重新命名後缺少欄位：{}\n\n新版目前欄位：{}\n\n"
            "常見原因：新版欄位被重複對應，或欄位對應選錯。".format(
                missing_new_after_rename,
                list(new_df.columns),
            )
        )

    old_df = old_df[old_needed].copy()
    new_df = new_df[new_needed].copy()

    return old_df, new_df, compare_key, compare_cols


# =========================================================
# 側邊欄設定
# =========================================================

with st.sidebar:
    st.header("比對設定")

    explode_refdes = st.checkbox(
        "將 Key 欄位用逗號拆成多列",
        value=True,
        help="適合 Ref Des 欄位，例如 C1,C2,C3。",
    )

    case_sensitive = st.checkbox(
        "大小寫視為不同",
        value=False,
    )

    dni_compare = st.checkbox(
        "DNI比較",
        value=False,
        help=(
            "開啟後，若指定的 BOM_IGNORE / DNI 欄位內容含有 DNI，"
            "會依 DNI 規則判斷 Add / Del / Changed / Unchanged。"
        ),
    )

    remove_blank_key = st.checkbox(
        "移除空白 Key、?、nan",
        value=True,
        help="建議開啟。會清除Ref中的 ?、？、N/A、nan、None、null。",
    )

    question_as_blank = st.checkbox(
        "將 ?、？、N/A、nan 視為空白",
        value=True,
        help="建議開啟。會清除所有欄位中的 ?、？、N/A、nan、None、null。",
    )

    show_unchanged = st.checkbox(
        "預覽畫面顯示 Unchanged",
        value=False,
    )


# =========================================================
# 上傳檔案
# =========================================================

col_upload_1, col_upload_2 = st.columns(2)

with col_upload_1:
    old_file = st.file_uploader(
        "上傳舊版 BOM",
        type=["csv", "xls", "xlsx", "xlsm"],
        key="old_file",
    )

with col_upload_2:
    new_file = st.file_uploader(
        "上傳新版 BOM",
        type=["csv", "xls", "xlsx", "xlsm"],
        key="new_file",
    )

if old_file is None or new_file is None:
    st.info("請先上傳舊版 BOM 與新版 BOM。")
    st.stop()


# =========================================================
# Sheet 與 Header Row 選擇
# =========================================================

try:
    old_sheets = read_input_sheets(old_file)
    new_sheets = read_input_sheets(new_file)
except Exception as e:
    show_friendly_error(
        title="讀取檔案失敗",
        message="請確認上傳的是有效的 csv、xls、xlsx、xlsm檔案。",
        detail=e,
    )

st.subheader("步驟 1：選擇 Sheet 與 Header Row")

col_sheet_1, col_sheet_2 = st.columns(2)

with col_sheet_1:
    old_sheet = st.selectbox(
        "舊版 BOM Sheet",
        old_sheets,
        key="old_sheet",
    )

    old_raw = read_table_raw(old_file, old_sheet)
    old_detected_header = detect_header_row(old_raw)

    old_header_excel_row = st.number_input(
        "舊版 Header Row，Excel / CSV 列號",
        min_value=1,
        max_value=max(1, len(old_raw)),
        value=int(old_detected_header + 1),
        step=1,
        key="old_header_row",
    )

with col_sheet_2:
    new_sheet = st.selectbox(
        "新版 BOM Sheet",
        new_sheets,
        key="new_sheet",
    )

    new_raw = read_table_raw(new_file, new_sheet)
    new_detected_header = detect_header_row(new_raw)

    new_header_excel_row = st.number_input(
        "新版 Header Row，Excel / CSV 列號",
        min_value=1,
        max_value=max(1, len(new_raw)),
        value=int(new_detected_header + 1),
        step=1,
        key="new_header_row",
    )

try:
    old_df_raw = load_bom_dataframe(
        old_file,
        old_sheet,
        int(old_header_excel_row - 1),
    )

    new_df_raw = load_bom_dataframe(
        new_file,
        new_sheet,
        int(new_header_excel_row - 1),
    )

    if question_as_blank:
        old_df_raw = clean_dataframe_for_output(old_df_raw)
        new_df_raw = clean_dataframe_for_output(new_df_raw)

except Exception as e:
    show_friendly_error(
        title="載入 BOM 資料失敗",
        message=(
            "請確認檔案、Sheet 與 Header Row 是否正確。\n\n"
            "如果欄位名稱沒有正確讀取，請手動調整 Header Row。"
        ),
        detail=e,
    )


# =========================================================
# 欄位選擇
# =========================================================

st.subheader("步驟 2：選擇 Key 欄位與要比對的欄位")
st.markdown("🔎 **說明：** 請選擇元件 location")


old_cols = list(old_df_raw.columns)
new_cols = list(new_df_raw.columns)

if len(old_cols) == 0 or len(new_cols) == 0:
    show_friendly_error(
        title="找不到欄位",
        message="請確認檔案、Sheet 與 Header Row 是否正確。",
    )

key_candidates = [
    "Ref Des",
    "REF DES",
    "Reference Designator",
    "Designator",
    "Location",
]

part_candidates = [
    "PART_NUMBER",
    "PART NUMBER",
    "Part Number",
    "PN",
    "P/N",
    "Item Number",
]

mfg_part_candidates = [
    "MFG_PART_NUMBER",
    "MFG PART NUMBER",
    "Mfr. Part Number",
    "Mfr Part Number",
    "Manufacturer Part Number",
    "MPN",
]

mfg_name_candidates = [
    "MFG",
    "Mfr. Name",
    "Mfr Name",
    "Manufacturer",
    "Manufacturer Name",
]

bom_ignore_candidates = [
    "BOM_IGNORE",
    "BOM IGNORE",
    "DNI",
    "Mount",
]

old_guess_key = guess_column(old_cols, key_candidates)
new_guess_key = guess_column(new_cols, key_candidates)

col_key_1, col_key_2 = st.columns(2)

with col_key_1:
    old_key_col = st.selectbox(
        "舊版 Key 欄位",
        old_cols,
        index=old_cols.index(old_guess_key) if old_guess_key in old_cols else 0,
    )

with col_key_2:
    new_key_col = st.selectbox(
        "新版 Key 欄位",
        new_cols,
        index=new_cols.index(new_guess_key) if new_guess_key in new_cols else 0,
    )

st.markdown("### 欄位對應")
st.caption("如果兩份 BOM 欄位名稱不同，可以在這裡指定對應關係。")

default_old_compare_guesses = [
    guess_column(old_cols, part_candidates),
    guess_column(old_cols, mfg_name_candidates),
    guess_column(old_cols, mfg_part_candidates),
    guess_column(old_cols, bom_ignore_candidates),
]

default_old_compare_guesses = [
    c for c in default_old_compare_guesses
    if c is not None and c != old_key_col
]

old_compare_options = [
    c for c in old_cols
    if c != old_key_col
]

old_compare_cols = st.multiselect(
    "選擇舊版要比對的欄位",
    options=old_compare_options,
    default=default_old_compare_guesses,
)

if len(old_compare_cols) == 0:
    st.warning("請至少選擇一個要比對的欄位。")
    st.stop()

new_compare_options = [
    c for c in new_cols
    if c != new_key_col
]

column_mapping = {}

for old_col in old_compare_cols:
    guess_new = None

    if old_col in new_compare_options:
        guess_new = old_col
    else:
        old_upper = old_col.upper()

        if old_upper in ["PN", "P/N"] or "PART" in old_upper:
            if (
                "MFR" in old_upper
                or "MFG" in old_upper
                or "MANUFACTURER" in old_upper
            ):
                guess_new = guess_column(new_compare_options, mfg_part_candidates)
            else:
                guess_new = guess_column(new_compare_options, part_candidates)

        elif (
            "MFR" in old_upper
            or "MFG" in old_upper
            or "MANUFACTURER" in old_upper
        ):
            guess_new = guess_column(new_compare_options, mfg_name_candidates)

        elif "IGNORE" in old_upper or "DNI" in old_upper:
            guess_new = guess_column(new_compare_options, bom_ignore_candidates)

    default_index = (
        new_compare_options.index(guess_new)
        if guess_new in new_compare_options
        else 0
    )

    selected_new_col = st.selectbox(
        "將舊版欄位 [" + old_col + "] 對應到新版欄位",
        options=new_compare_options,
        index=default_index,
        key="map_" + old_col,
    )

    column_mapping[old_col] = selected_new_col


# =========================================================
# 欄位對應檢查
# =========================================================

mapping_preview_df = pd.DataFrame(
    [
        {
            "舊版欄位": old_col,
            "新版欄位": new_col,
        }
        for old_col, new_col in column_mapping.items()
    ]
)

with st.expander("目前欄位對應檢查", expanded=True):
    st.dataframe(mapping_preview_df, use_container_width=True)

mapped_new_cols = list(column_mapping.values())

duplicated_new_cols = sorted(
    {
        col for col in mapped_new_cols
        if mapped_new_cols.count(col) > 1
    }
)

if duplicated_new_cols:
    duplicate_detail = []

    for new_col in duplicated_new_cols:
        related_old_cols = [
            old_col
            for old_col, mapped_new_col in column_mapping.items()
            if mapped_new_col == new_col
        ]

        duplicate_detail.append(
            {
                "重複的新版欄位": new_col,
                "被哪些舊版欄位使用": ", ".join(related_old_cols),
            }
        )

    duplicate_df = pd.DataFrame(duplicate_detail)

    st.error("欄位對應設定錯誤")
    st.warning(
        "同一個新版欄位不可被多個舊版欄位重複使用。\n\n"
        "請檢查下方重複對應表，重新選擇正確的新版欄位。"
    )
    st.dataframe(duplicate_df, use_container_width=True)

    st.info(
        "例如：舊版 MFG 不應該對應到新版 PART_NUMBER；"
        "請改成新版的 MFG、Mfr. Name、Manufacturer 或相對應欄位。"
    )

    st.stop()


# =========================================================
# DNI 欄位設定
# =========================================================

bom_ignore_col = None
extra_keep_cols = []

if dni_compare:
    st.markdown("### DNI 欄位設定")
    st.caption("如果只有其中一份 BOM 有 DNI 欄位，另一份請選「不使用」。")

    old_dni_candidates = [
        col for col in old_df_raw.columns
        if is_possible_dni_column(col)
    ]

    new_dni_candidates = [
        col for col in new_df_raw.columns
        if is_possible_dni_column(col)
    ]

    old_dni_options = ["不使用"] + old_dni_candidates
    new_dni_options = ["不使用"] + new_dni_candidates

    dni_col_1, dni_col_2 = st.columns(2)

    with dni_col_1:
        old_dni_col = st.selectbox(
            "舊版 BOM DNI 判斷欄位",
            options=old_dni_options,
            index=1 if len(old_dni_options) > 1 else 0,
            help="如果舊版 BOM 沒有 DNI 欄位，請選不使用。",
        )

    with dni_col_2:
        new_dni_col = st.selectbox(
            "新版 BOM DNI 判斷欄位",
            options=new_dni_options,
            index=1 if len(new_dni_options) > 1 else 0,
            help="如果新版 BOM 沒有 DNI 欄位，請選不使用。",
        )

    old_df_raw[DNI_COMPARE_COL] = ""
    new_df_raw[DNI_COMPARE_COL] = ""

    if old_dni_col != "不使用":
        old_df_raw[DNI_COMPARE_COL] = old_df_raw[old_dni_col].apply(
            lambda x: clean_value(x, question_as_blank=True)
        )

    if new_dni_col != "不使用":
        new_df_raw[DNI_COMPARE_COL] = new_df_raw[new_dni_col].apply(
            lambda x: clean_value(x, question_as_blank=True)
        )

    bom_ignore_col = DNI_COMPARE_COL
    extra_keep_cols = [DNI_COMPARE_COL]


# =========================================================
# 欄位標準化
# =========================================================

try:
    old_df_norm, new_df_norm, compare_key_col, compare_cols = rename_for_compare(
        old_df_raw,
        new_df_raw,
        old_key_col,
        new_key_col,
        column_mapping,
        extra_keep_cols=extra_keep_cols,
    )
except Exception as e:
    show_friendly_error(
        title="欄位對應設定錯誤",
        message=(
            "系統無法建立比對欄位，請檢查以下項目：\n\n"
            "1. 舊版欄位是否對應到正確的新版欄位。\n"
            "2. 同一個新版欄位是否被重複使用。\n"
            "3. Header Row 是否選錯，導致欄位名稱讀取錯誤。\n"
            "4. PART_NUMBER、MFG、MFG_PART_NUMBER、BOM_IGNORE 是否各自對到正確欄位。"
        ),
        detail=e,
    )


# =========================================================
# 原始資料預覽
# =========================================================

with st.expander("原始資料預覽", expanded=False):
    preview_col_1, preview_col_2 = st.columns(2)

    with preview_col_1:
        st.write("舊版 BOM 預覽")
        st.dataframe(old_df_raw.head(30), use_container_width=True)

    with preview_col_2:
        st.write("新版 BOM 預覽")
        st.dataframe(new_df_raw.head(30), use_container_width=True)


# =========================================================
# 執行比對
# =========================================================

st.subheader("步驟 3：執行比對")

run_button = st.button(
    "開始比對",
    type="primary",
    use_container_width=True,
)

if not run_button:
    st.stop()

try:
    old_prepared = prepare_bom(
        old_df_norm,
        key_col=compare_key_col,
        compare_cols=compare_cols,
        explode_key=explode_refdes,
        remove_blank_key=remove_blank_key,
        extra_cols=extra_keep_cols,
    )

    new_prepared = prepare_bom(
        new_df_norm,
        key_col=compare_key_col,
        compare_cols=compare_cols,
        explode_key=explode_refdes,
        remove_blank_key=remove_blank_key,
        extra_cols=extra_keep_cols,
    )

    result_df, detail_df = compare_bom(
        old_prepared,
        new_prepared,
        key_col=compare_key_col,
        compare_cols=compare_cols,
        case_sensitive=case_sensitive,
        dni_compare=dni_compare,
        bom_ignore_col=bom_ignore_col,
    )

    if old_key_col == new_key_col:
        display_key_name = old_key_col
    else:
        display_key_name = old_key_col + " / " + new_key_col

    result_df = result_df.rename(columns={compare_key_col: display_key_name})
    detail_df = detail_df.rename(columns={compare_key_col: display_key_name})

    result_df = clean_dataframe_for_output(result_df)
    detail_df = clean_dataframe_for_output(detail_df)

    change_list_df = make_change_list(result_df)
    change_list_df = clean_dataframe_for_output(change_list_df)

    summary_df = make_summary(
        result_df,
        old_file_name=old_file.name,
        new_file_name=new_file.name,
        old_sheet_name=old_sheet,
        new_sheet_name=new_sheet,
    )

except Exception as e:
    show_friendly_error(
        title="比對失敗",
        message=(
            "BOM 比對過程中發生錯誤。\n\n"
            "請確認 Key 欄位、欄位對應、以及 BOM 資料內容是否正確。"
        ),
        detail=e,
    )


# =========================================================
# 結果摘要
# =========================================================

st.subheader("步驟 4：比對結果摘要")

status_counts = result_df["狀態"].value_counts()

metric_col_1, metric_col_2, metric_col_3, metric_col_4, metric_col_5 = st.columns(5)

metric_col_1.metric("總筆數", len(result_df))
metric_col_2.metric("Changed", int(status_counts.get("Changed", 0)))
metric_col_3.metric("New Add", int(status_counts.get("New Add", 0)))
metric_col_4.metric("Del", int(status_counts.get("Del", 0)))
metric_col_5.metric("Unchanged", int(status_counts.get("Unchanged", 0)))

if dni_compare:
    st.info("目前已啟用 DNI比較。支援只有舊版或只有新版有 DNI 欄位。")


# =========================================================
# 結果頁面
# =========================================================

st.subheader("步驟 5：比對結果")

tab_change_list, tab_compare_result, tab_detail = st.tabs(
    ["Change List", "Compare Result", "欄位差異 Detail"]
)

with tab_change_list:
    st.markdown("### 📋 Change List")
    st.caption("只顯示 Changed / New Add / Del，Unchanged 不會顯示。")

    change_status_options = (
        list(change_list_df["狀態"].dropna().unique())
        if not change_list_df.empty
        else []
    )

    if change_status_options:
        selected_change_status = st.multiselect(
            "Change List 狀態篩選",
            options=change_status_options,
            default=change_status_options,
            key="change_list_status_filter",
        )

        filtered_change_list_df = change_list_df.copy()

        if selected_change_status:
            filtered_change_list_df = filtered_change_list_df[
                filtered_change_list_df["狀態"].isin(selected_change_status)
            ].copy()
    else:
        filtered_change_list_df = change_list_df.copy()

    st.dataframe(
        filtered_change_list_df,
        use_container_width=True,
        height=500,
    )

with tab_compare_result:
    st.markdown("### 🔍 Compare Result")

    if show_unchanged:
        preview_df = result_df.copy()
    else:
        preview_df = result_df[result_df["狀態"] != "Unchanged"].copy()

    preview_df = clean_dataframe_for_output(preview_df)

    status_filter_options = list(preview_df["狀態"].dropna().unique())

    selected_status = st.multiselect(
        "Compare Result 狀態篩選",
        options=status_filter_options,
        default=status_filter_options,
        key="compare_result_status_filter",
    )

    if selected_status:
        preview_df = preview_df[preview_df["狀態"].isin(selected_status)].copy()

    st.dataframe(
        preview_df,
        use_container_width=True,
        height=500,
    )

with tab_detail:
    st.markdown("### 🧾 欄位差異 Detail")

    detail_df = clean_dataframe_for_output(detail_df)

    st.dataframe(
        detail_df,
        use_container_width=True,
        height=500,
    )


# =========================================================
# 下載 Excel
# =========================================================

try:
    excel_bytes = style_excel_output(
        result_df,
        change_list_df,
        detail_df,
        summary_df,
    )
except Exception as e:
    show_friendly_error(
        title="Excel 輸出失敗",
        message="產生 Excel 結果檔時發生錯誤，請檢查資料內容是否有異常。",
        detail=e,
    )

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

def safe_filename_part(name, max_len=35):
    name = str(name).rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    if "." in name:
        name = name.rsplit(".", 1)[0]
    name = re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
    return name[:max_len] if len(name) > max_len else name

output_filename = (
    "BOM_Compare_Result_"
    + timestamp
    + ".xlsx"
)

st.download_button(
    label="下載 Excel 比對結果",
    data=excel_bytes,
    file_name=output_filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

st.success("BOM 比對完成。")
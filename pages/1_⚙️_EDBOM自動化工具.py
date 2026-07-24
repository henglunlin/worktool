import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import re
import os
import shutil
import time
import math
import traceback
import copy
from collections import defaultdict 

# =====================================================
# ✅ Streamlit UI 設定與 Session State 初始化
# =====================================================
st.set_page_config(page_title="EDBOM 自動化工具", page_icon="🔧", layout="wide")
st.title("🔧 EDBOM 自動化工具")
st.markdown("將 `EDlist` 與 `BOM` 比對，如有缺料自動從 `RLClist` 補齊，並告訴自己要買日月光。。")

# 新增：折疊式使用說明與檔案用途說明
with st.expander("📖 點此查看工具使用說明與檔案用途", expanded=False):
    st.markdown("""
    ### 📝 使用說明
    1. **上傳檔案**：於左側側邊欄依序上傳所需的檔案（EDlist, BOM, temp樣板, RLClist）。
    2. **解析欄位**：上傳完畢後，點擊左側 **「🔍 讀取檔案與解析欄位」** 按鈕。
    3. **確認對應**：在下方的 **「⚙️ 欄位對應設定」** 區塊，檢查系統自動抓取的欄位是否正確（若檔案中無該欄位，請保持 `[忽略]`）。
    4. **開始執行**：滑至最下方點擊 **「🚀 開始執行自動化流程」**，等待綠色成功訊息後即可下載成果。

    ---
    ### 📂 輸入檔案用途說明
    * **EDlist.xlsx (變更清單)**：記載本次需要異動的料件，包含變更動作 (Add/Del)、料號 (PN) 與位置 (Ref Des)。
    * **BOM.xls / .xlsx (原始物料清單)**：原始的 BOM 表，提供目前系統中的料件狀態與現有數量。
    * **temp.xlsx (輸出樣板)**：這是一個「空表」，系統會依照此檔案的「表頭欄位」與「格式樣式（字體/框線/顏色）」來產出最終的 `EDBOM_updated.xlsx`。
    * **RLClist.xlsx (總料表庫)**：當 EDlist 中有新增料件，但原始 BOM 找不到時，系統會自動來此檔案撈取料號的詳細屬性（如 Description、廠牌等）進行補料。

    ---
    ### 📥 輸出成果檔案用途說明
    * **EDBOM_updated.xlsx (最終完成表)**：最終EC_BOM 表。
    * **newpart.xlsx (缺料原始清單)**：記錄初步比對時，有出現在 EDlist 裡，但在原始 BOM 中找不到的料號清單。
    * **qty.xlsx (數量異動彙總表)**：包含「AddDel明細」與「Qty彙總」兩個分頁，方便核對各料件原始數量、增減變化與最終數量的計算過程。
    * **RLC_result.xlsx (補料成功清單)**：記錄系統自動從 `RLClist.xlsx` 中成功找到並補齊屬性資料的新增料件。
    * **RLC_NotFound.xlsx (完全找不到的料件)**：⚠️ **需人工處理**。記錄既不在原始 BOM 裡，也無法從 RLClist 中找到的料件，代表系統無法取得該料號的詳細資訊。
    """)

if "run_complete" not in st.session_state:
    st.session_state.run_complete = False
    st.session_state.has_newpart = False
    st.session_state.show_balloons = False

if "debug_log_text" not in st.session_state:
    st.session_state.debug_log_text = ""

if "files_loaded" not in st.session_state:
    st.session_state.files_loaded = False

if "auto_bom_row_val" not in st.session_state:
    st.session_state.auto_bom_row_val = 0

# =====================================================
# ✅ 工具與匯出函式 (快取與效能優化)
# =====================================================
def export_excel_with_calibri(df, filename):
    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        calibri_format = workbook.add_format({'font_name': 'Calibri', 'font_size': 11})
        worksheet.set_column(0, len(df.columns) - 1, None, calibri_format)

def normalize_find_num(v):
    if v is None: return ""
    if isinstance(v, float) and math.isnan(v): return ""
    s = str(v).strip()
    if s == "" or s.lower() == "nan": return ""
    if s.lower() == "new": return "NEW"
    try:
        f = float(s)
        if abs(f - round(f)) < 1e-9: return str(int(round(f)))
        return str(f)
    except:
        return s.upper()

def sort_key_for_find_num(v):
    key = normalize_find_num(v)
    if key == "NEW": return (2, float("inf"), key)
    try: return (0, float(key), key)
    except: return (1, float("inf"), key)

def expand_ref_des_token(token):
    """將單一 Ref Des token 展開，例如 'WC11361-WC11372' -> ['WC11361', 'WC11362', ..., 'WC11372']
    如果不是合法的區間格式（前綴不一致、結尾小於開頭等），則原樣傳回，避免誤判。"""
    token = token.strip()
    if not token or "-" not in token:
        return [token] if token else []
    m = re.match(r'^([A-Za-z]+)(\d+)\s*-\s*([A-Za-z]*)(\d+)$', token)
    if not m:
        return [token]
    prefix, start_num, prefix2, end_num = m.groups()
    if prefix2 and prefix2 != prefix:
        return [token]
    width = len(start_num)
    start, end = int(start_num), int(end_num)
    if end < start or (end - start) > 5000:  # 防呆：避免異常區間造成大量展開
        return [token]
    return [f"{prefix}{str(i).zfill(width)}" for i in range(start, end + 1)]

def expand_ref_des(ref_str):
    """展開整串 Ref Des（可能包含逗號分隔的多個 token，其中部分可能是 'X1-X5' 這種壓縮區間格式）。
    有些 BOM 匯出格式（例如 Agile）會把 Ref Des 壓縮成區間，若不展開，逐一比對 EDlist 時會抓不到料件。"""
    if ref_str is None: return ""
    s = str(ref_str).strip()
    if s == "" or s.lower() == "nan": return ""
    tokens = [t.strip() for t in s.replace(" ", "").split(",") if t.strip()]
    expanded = []
    for t in tokens:
        expanded.extend(expand_ref_des_token(t))
    # 去重並保留原始出現順序
    seen = set()
    result = []
    for r in expanded:
        if r not in seen:
            seen.add(r)
            result.append(r)
    return ",".join(result)

class RunLogger:
    """統一收集執行過程中的詳細日誌，方便匯出除錯。"""
    def __init__(self, log_dir):
        self.log_dir = log_dir
        self.lines = []
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)

    def log(self, msg, level="INFO"):
        ts = time.strftime("%H:%M:%S")
        line = f"[{ts}] [{level}] {msg}"
        self.lines.append(line)
        return line

    def info(self, msg): return self.log(msg, "INFO")
    def warn(self, msg): return self.log(msg, "WARN")
    def error(self, msg): return self.log(msg, "ERROR")

    def text(self):
        return "\n".join(self.lines)

    def save(self, filename="run.log"):
        try:
            with open(os.path.join(self.log_dir, filename), "w", encoding="utf-8") as f:
                f.write(self.text())
        except Exception:
            pass

def get_match_index(columns, keywords):
    """加入 [忽略] 的智慧比對：如果找不到，就預設回傳 0 (即 [忽略])"""
    for kw in keywords:
        for idx, col in enumerate(columns):
            if col == "[忽略]": continue
            if kw.lower() == str(col).lower().strip(): return idx
    for kw in keywords:
        for idx, col in enumerate(columns):
            if col == "[忽略]": continue
            if str(col).lower().strip().startswith(kw.lower()): return idx
    for kw in keywords:
        for idx, col in enumerate(columns):
            if col == "[忽略]": continue
            if kw.lower() in str(col).lower(): return idx
    return 0 

def smart_read_excel_header(filepath, header_row=None, nrows=None, logger=None):
    errors = []
    for eng in ["openpyxl", "xlrd"]:
        try:
            df = pd.read_excel(filepath, header=header_row, nrows=nrows, engine=eng)
            if logger: logger.info(f"讀取 '{filepath}' 成功 (engine={eng}, header_row={header_row}, shape={df.shape})")
            return df
        except Exception as e:
            errors.append(f"{eng}: {e}")
            continue
    try:
        df = pd.read_excel(filepath, header=header_row, nrows=nrows)
        if logger: logger.info(f"讀取 '{filepath}' 成功 (engine=預設, header_row={header_row}, shape={df.shape})")
        return df
    except Exception as e:
        if logger:
            logger.error(f"讀取 '{filepath}' 失敗，所有 engine 皆失敗：{' | '.join(errors)} | 預設: {e}")
        raise

def find_bom_header(filepath):
    try:
        df_temp = smart_read_excel_header(filepath, header_row=None, nrows=40)
        for idx, row in df_temp.iterrows():
            row_str = " ".join(row.fillna("").astype(str).str.lower().tolist())
            if ("find num" in row_str and "item number" in row_str) or ("part_number" in row_str) or ("ref des" in row_str and "qty" in row_str):
                return idx
    except Exception: 
        pass
    return 0 

def get_excel_columns(filepath, header_row=0):
    if not os.path.exists(filepath): return []
    try:
        df = smart_read_excel_header(filepath, header_row=header_row, nrows=0)
        return df.columns.str.strip().tolist()
    except:
        return []

def save_intermediate_logs(log_dir, df_dict):
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    for filename, df in df_dict.items():
        if df is None:
            continue
        if isinstance(df, dict):
            # 多分頁輸出：df 為 {分頁名稱: DataFrame}
            sheets = {name: d for name, d in df.items() if d is not None and not d.empty}
            if not sheets:
                continue
            with pd.ExcelWriter(os.path.join(log_dir, filename), engine="openpyxl") as writer:
                for sheet_name, sheet_df in sheets.items():
                    sheet_df.to_excel(writer, index=False, sheet_name=str(sheet_name)[:31])
        else:
            if not df.empty:
                df.to_excel(os.path.join(log_dir, filename), index=False)

def get_col_data(df, col_name, default_val):
    """安全獲取欄位資料：如果是 [忽略] 或找不到，就給預設值"""
    if col_name != "[忽略]" and col_name in df.columns:
        return df[col_name]
    if isinstance(default_val, list) and len(default_val) == len(df):
        return default_val
    return [default_val] * len(df)

# =====================================================
# ✅ 側邊欄：檔案上傳與解析
# =====================================================
st.sidebar.header("📂 檔案上傳與解析")
st.sidebar.info("💡 上傳檔案後，請點擊下方按鈕來讀取與解析表頭欄位。")

upload_ed = st.sidebar.file_uploader("1. 上傳 EDlist.xlsx (可選)", type=["xlsx"])
upload_bom = st.sidebar.file_uploader("2. 上傳 BOM (可選)", type=["xls", "xlsx"])
upload_temp = st.sidebar.file_uploader("3. 上傳 temp.xlsx [樣板] (可選)", type=["xlsx"])
upload_rlc = st.sidebar.file_uploader("4. 上傳 RLClist.xlsx (可選)", type=["xlsx"])

file_mapping = {"EDlist.xlsx": upload_ed, "BOM.xls": upload_bom, "temp.xlsx": upload_temp, "RLClist.xlsx": upload_rlc}
for filename, uploaded_file in file_mapping.items():
    if uploaded_file is not None:
        save_name = "BOM.xls" if filename == "BOM.xls" and uploaded_file.name.endswith(".xlsx") else filename
        with open(save_name, "wb") as f: f.write(uploaded_file.getbuffer())

bom_filename = "BOM.xls"
if not os.path.exists(bom_filename):
    for f in os.listdir("."):
        if "bom" in f.lower() and f.endswith((".xls", ".xlsx")):
            bom_filename = f
            break

if st.sidebar.button("🔍 讀取檔案與解析欄位", use_container_width=True):
    missing_files = [f for f in ["EDlist.xlsx", "temp.xlsx", "RLClist.xlsx"] if not os.path.exists(f)]
    if missing_files or not os.path.exists(bom_filename):
        st.sidebar.error("⚠️ 缺少必要檔案，請確認已上傳所有檔案！")
        st.session_state.files_loaded = False
    else:
        st.session_state.auto_bom_row_val = int(find_bom_header(bom_filename))
        st.session_state.files_loaded = True
        st.sidebar.success("✅ 檔案讀取成功！")

if not st.session_state.files_loaded:
    missing_initial = [f for f in ["EDlist.xlsx", "temp.xlsx", "RLClist.xlsx", bom_filename] if not os.path.exists(f)]
    if missing_initial:
        st.warning(f"⚠️ 請在左側上傳或確認必要檔案已備妥，並點擊側邊欄的 **「🔍 讀取檔案與解析欄位」** 按鈕開始。")
        st.stop()
    else:
        st.session_state.auto_bom_row_val = int(find_bom_header(bom_filename))
        st.session_state.files_loaded = True

st.sidebar.markdown("---")
st.sidebar.header("⚙️ 進階設定")
bom_header_row = st.sidebar.number_input(
    "BOM 表標題列位置", 
    min_value=0, max_value=50, 
    value=st.session_state.auto_bom_row_val
)

# 加上 [忽略] 選項
ed_cols = ["[忽略]"] + get_excel_columns("EDlist.xlsx")
bom_cols = ["[忽略]"] + get_excel_columns(bom_filename, header_row=bom_header_row)
temp_cols = get_excel_columns("temp.xlsx") # temp 樣板不需要忽略
rlc_cols = ["[忽略]"] + get_excel_columns("RLClist.xlsx")

if not temp_cols:
    st.error(f"❌ 讀取欄位失敗，請確認檔案格式或 BOM 標題列位置 (目前設為 {bom_header_row}) 是否正確後，重新點擊側邊欄的讀取按鈕。")
    st.stop()

# =====================================================
# ✅ 欄位對應設定 UI
# =====================================================
st.subheader("⚙️ 欄位對應設定 (Column Mapping)")
st.markdown("#### 📦 BOM 欄位")
b_cols = st.columns(7)
bom_map_find = b_cols[0].selectbox("BOM: Find Num", bom_cols, index=get_match_index(bom_cols, ["find", "find num", "index"]))
bom_map_item = b_cols[1].selectbox("BOM: Item Number", bom_cols, index=get_match_index(bom_cols, ["item number", "part_number", "pn"]))
bom_map_desc = b_cols[2].selectbox("BOM: Item Description", bom_cols, index=get_match_index(bom_cols, ["description", "agile_description", "desc"]))
bom_map_qty  = b_cols[3].selectbox("BOM: Qty", bom_cols, index=get_match_index(bom_cols, ["qty", "數量"]))
bom_map_ref  = b_cols[4].selectbox("BOM: Ref Des", bom_cols, index=get_match_index(bom_cols, ["ref des", "ref"]))
bom_map_mfr  = b_cols[5].selectbox("BOM: Mfr. Name", bom_cols, index=get_match_index(bom_cols, ["mfr. name", "mfg", "manufacturer"]))
bom_map_mpn  = b_cols[6].selectbox("BOM: Mfr. Part Number", bom_cols, index=get_match_index(bom_cols, ["mfr. part number", "mfg_part_number", "part number"]))
st.markdown("---")

st.markdown("#### 📝 EDlist 欄位")
e_cols = st.columns(3)
ed_map_change = e_cols[0].selectbox("變更動作 (Change)", ed_cols, index=get_match_index(ed_cols, ["change", "action", "動作"]))
ed_map_pn     = e_cols[1].selectbox("料號 (PN)", ed_cols, index=get_match_index(ed_cols, ["pn", "item", "料號", "number"]))
ed_map_ref    = e_cols[2].selectbox("位置 (Ref Des)", ed_cols, index=get_match_index(ed_cols, ["ref", "位置"]))
st.markdown("---")

st.markdown("#### 🔍 RLClist 欄位")
r_cols = st.columns(4)
rlc_map_item = r_cols[0].selectbox("RLC: Item Number", rlc_cols, index=get_match_index(rlc_cols, ["item", "pn"]))
rlc_map_desc = r_cols[1].selectbox("RLC: Item Description", rlc_cols, index=get_match_index(rlc_cols, ["desc"]))
rlc_map_mfr  = r_cols[2].selectbox("RLC: Mfr. Name", rlc_cols, index=get_match_index(rlc_cols, ["mfr. name", "mfr name", "manufacturer"]))
rlc_map_mpn  = r_cols[3].selectbox("RLC: Mfr. Part Number", rlc_cols, index=get_match_index(rlc_cols, ["part number", "mpn"]))
st.markdown("---")

# =====================================================
# ✅ 核心邏輯模組化 (防呆處理 [忽略] 與預設值)
# =====================================================
def step1_match_ed_bom(ed_df_raw, bom_df_raw, maps, logger=None):
    # 安全獲取 EDlist 欄位，若為 [忽略] 則預設為空白
    ed_df = pd.DataFrame({
        "change": get_col_data(ed_df_raw, maps["ed_change"], ""),
        "PN": get_col_data(ed_df_raw, maps["ed_pn"], ""),
        "Ref Des": get_col_data(ed_df_raw, maps["ed_ref"], "")
    })
    
    # 安全獲取 BOM 欄位，若 Find Num 被忽略，則自動補上流水號
    if maps["bom_find"] == "[忽略]" and logger:
        logger.warn("BOM 欄位對應中『Find Num』為 [忽略] 或找不到對應欄位，將自動以 NEW_0, NEW_1... 流水號補上。")
    default_find = [f"NEW_{i}" for i in range(len(bom_df_raw))]
    bom_df = pd.DataFrame({
        "Find Num": get_col_data(bom_df_raw, maps["bom_find"], default_find),
        "Item Number": get_col_data(bom_df_raw, maps["bom_item"], ""),
        "Item Description": get_col_data(bom_df_raw, maps["bom_desc"], ""), 
        "Qty": get_col_data(bom_df_raw, maps["bom_qty"], 1),
        "Ref Des": get_col_data(bom_df_raw, maps["bom_ref"], ""), 
        "Mfr. Name": get_col_data(bom_df_raw, maps["bom_mfr"], ""),
        "Mfr. Part Number": get_col_data(bom_df_raw, maps["bom_mpn"], "")
    })

    ed_df["PN"] = ed_df["PN"].astype(str).str.strip()
    ed_df["Ref Des"] = ed_df["Ref Des"].astype(str).str.strip()
    bom_df["Item Number"] = bom_df["Item Number"].astype(str).str.strip()

    # 有些 BOM 匯出格式（例如 Agile 的複合 Ref Des 欄位）會把連續位置壓縮成區間，例如 "WC11361-WC11372"，
    # 若不展開成個別位置，逐一比對 EDlist 時會抓不到料件，因此統一展開後再比對與輸出。
    bom_df["Ref Des"] = bom_df["Ref Des"].astype(str).apply(expand_ref_des)

    ed_item = ed_df["PN"].dropna()
    ref_set = set([r.strip() for r in ed_df["Ref Des"].dropna().str.split(",").explode().tolist() if r.strip()])
    
    cond_pn = bom_df["Item Number"].isin(ed_item)
    cond_ref = bom_df["Ref Des"].apply(lambda x: not set(str(x).replace(" ", "").split(",")).isdisjoint(ref_set))

    matched_bom = bom_df[cond_pn | cond_ref][["Find Num", "Item Number", "Item Description", "Qty", "Ref Des", "Mfr. Name", "Mfr. Part Number"]].copy()
    missing_df = ed_df[~ed_df["PN"].isin(bom_df["Item Number"])][["PN", "Ref Des"]].copy()
    missing_df.rename(columns={"PN": "Missing Item Number"}, inplace=True)

    if logger:
        logger.info(f"EDlist 共 {len(ed_df)} 筆；BOM 共 {len(bom_df)} 筆。")
        logger.info(f"透過料號(PN)比對命中 {int(cond_pn.sum())} 筆；透過 Ref Des 比對命中 {int(cond_ref.sum())} 筆；聯集共 {len(matched_bom)} 筆。")
        if matched_bom.empty:
            logger.warn("Step1 比對後 matched_bom 為 0 筆！請確認 BOM/EDlist 的欄位對應（Item Number / Ref Des）是否正確選取，或 BOM 是否為正確的檔案版本。")
        logger.info(f"EDlist 中有 {len(missing_df)} 筆料號在 BOM 中找不到，將嘗試以 RLClist 補齊。" if len(missing_df) else "EDlist 所有料號皆已在 BOM 中找到，略過 RLC 補料流程。")

    return matched_bom, missing_df, ed_df

def step1_5_rlc_search(missing_df, df_rlc_raw, matched_bom, maps, logger=None):
    # 安全獲取 RLC 欄位
    df_rlc = pd.DataFrame({
        "Item Number": get_col_data(df_rlc_raw, maps["rlc_item"], ""), 
        "Item Description": get_col_data(df_rlc_raw, maps["rlc_desc"], ""),
        "Mfr. Name": get_col_data(df_rlc_raw, maps["rlc_mfr"], ""), 
        "Mfr. Part Number": get_col_data(df_rlc_raw, maps["rlc_mpn"], "")
    })

    df_new = missing_df.rename(columns={"Missing Item Number": "Item Number"}).copy()
    df_new["Item Number"] = df_new["Item Number"].astype(str).str.strip()
    df_new["Ref Des"] = df_new["Ref Des"].astype(str).str.replace(" ", "", regex=False)
    
    pn_to_ref = df_new.groupby("Item Number")["Ref Des"].apply(lambda x: ",".join(sorted(x))).reset_index()
    df_rlc["Item Number"] = df_rlc["Item Number"].astype(str).str.strip()

    # 防呆：RLClist 若有重複料號，merge 時會造成資料列數暴增，因此僅保留第一筆
    dup_count = int(df_rlc["Item Number"].duplicated().sum())
    if dup_count > 0:
        if logger: logger.warn(f"RLClist 中發現 {dup_count} 筆重複的 Item Number，比對時僅保留第一筆對應資料。")
        df_rlc = df_rlc.drop_duplicates(subset="Item Number", keep="first")

    df_merge = pd.merge(pn_to_ref, df_rlc, on="Item Number", how="left")
    df_merge["Qty"] = df_merge["Ref Des"].apply(lambda ref: 0 if pd.isna(ref) or ref == "" else len(ref.split(",")))
    df_final = df_merge[["Item Number", "Item Description", "Qty", "Ref Des", "Mfr. Name", "Mfr. Part Number"]]
    
    found_df = df_final[df_final["Item Description"].notna()].copy()
    not_found_df = df_final[df_final["Item Description"].isna()].copy()

    if not found_df.empty:
        rlc_result = found_df.copy()
        rlc_result.insert(0, "Find Num", "New")
    else:
        rlc_result = pd.DataFrame(columns=["Find Num", "Item Number", "Item Description", "Qty", "Ref Des", "Mfr. Name", "Mfr. Part Number"])
        
    result_all = pd.concat([matched_bom, rlc_result], ignore_index=True)
    main_keys = set(matched_bom["Find Num"].apply(normalize_find_num))
    sup_keys = result_all["Find Num"].apply(normalize_find_num)
    
    df_sup_filtered = result_all[~sup_keys.isin(main_keys)].copy()
    df_all_sorted = pd.concat([matched_bom, df_sup_filtered], ignore_index=True)
    df_all_sorted["_sort_key"] = df_all_sorted["Find Num"].apply(sort_key_for_find_num)
    df_all_sorted = df_all_sorted.sort_values(by="_sort_key", kind="stable").drop(columns=["_sort_key"])

    if logger:
        logger.info(f"RLC 補料：缺料 {len(pn_to_ref)} 個料號中，{len(found_df)} 個在 RLClist 找到，{len(not_found_df)} 個完全找不到。")
        logger.info(f"合併後最終 BOM 筆數：{len(df_all_sorted)} 筆。")

    return df_all_sorted, not_found_df, found_df

def step2_calc_diff(res_df, ed_df, logger=None):
    adds_for_pn = defaultdict(set)
    dels_for_pn = defaultdict(set)
    all_adds = defaultdict(set)
    all_dels = defaultdict(set)
    
    for _, row in ed_df.iterrows():
        act = str(row["change"]).strip()
        pn = str(row["PN"]).strip()
        refs = [r.strip() for r in str(row["Ref Des"]).split(",") if r.strip()]
        for ref in refs:
            if act == "Add":
                adds_for_pn[pn].add(ref)
                all_adds[ref].add(pn)
            elif act == "Del":
                dels_for_pn[pn].add(ref)
                all_dels[ref].add(pn)

    applied_refs = set()

    def get_diff(row):
        pn = str(row["Item Number"]).strip()
        ref_list_set = set([r for r in str(row.get("Ref Des", "")).replace(" ", "").split(",") if r])
        add_list = adds_for_pn[pn].copy()
        del_list = dels_for_pn[pn].copy()
        
        for ref in ref_list_set:
            if ref in all_adds and pn not in all_adds[ref]:
                # 這個位置正被 Add 到別的料號（換料），視為從目前料號被移除
                del_list.add(ref)
            elif ref in all_dels and pn not in all_dels[ref] and ref not in add_list:
                # EDlist 記錄的 Del 是掛在「舊／不同料號」下（例如料號後來被替代料取代），
                # 但目前 BOM 這個位置實際上是掛在別的料號底下，仍要視為此列被刪除，
                # 否則這個 Ref Des 的異動會因為料號對不上而遺漏。
                del_list.add(ref)

        applied_refs.update(add_list)
        applied_refs.update(del_list)
        parts = []
        if add_list: parts.append("Add " + ",".join(sorted(add_list)))
        if del_list: parts.append("Del " + ",".join(sorted(del_list)))
        return ",".join(parts)

    res_df["變更差異"] = res_df.apply(get_diff, axis=1)

    if logger:
        all_ed_refs = set(all_adds.keys()) | set(all_dels.keys())
        unapplied = sorted(all_ed_refs - applied_refs)
        if unapplied:
            logger.warn(f"Step2：以下 EDlist 中的 Ref Des 未能對應到任何 BOM 列，請確認 BOM 中是否存在這些位置：{unapplied}")
        else:
            logger.info("Step2：EDlist 所有 Ref Des 異動皆已成功對應到 BOM 中的資料列。")

    def update_ref_des(row):
        ref_str = str(row.get("Ref Des", "")).replace(" ", "")
        if ref_str.lower() == "nan": ref_str = ""
        ref_set = set([r for r in ref_str.split(",") if r])
        change = str(row.get("變更差異", "")).strip()
        if change and change.lower() != "nan":
            add_match = re.search(r"Add\s+([A-Za-z0-9,]+?)(?=,?Del|$)", change)
            if add_match: ref_set.update([r.strip() for r in add_match.group(1).split(",") if r.strip()])
            del_match = re.search(r"Del\s+([A-Za-z0-9,]+)", change)
            if del_match: ref_set.difference_update([r.strip() for r in del_match.group(1).split(",") if r.strip()])
        return ",".join(sorted(ref_set))
    
    res_df["Ref Des"] = res_df.apply(update_ref_des, axis=1)
    return res_df

def step3_count_qty(res_df, logger=None):
    result_list = []
    skipped_new_rlc = 0
    for _, row in res_df.iterrows():
        pn, change, find_num = str(row.get("Item Number", "")).strip(), str(row.get("變更差異", "")).strip(), str(row.get("Find Num", "")).strip()
        # 注意：BOM 若無真實 Find Num 欄位，Step1 會用 "NEW_0", "NEW_1"... 當作預設流水號（這些仍是「原本就存在」的 BOM 料件，
        # Qty 一樣要套用 Add/Del）；只有 RLC 補料流程新增的料件，Find Num 才會是完全等於 "New" 的字串（這種料件的 Qty
        # 已經在 step1_5_rlc_search 依照 Ref Des 數量算好了，不能再重複加減，所以才要跳過）。
        # 過去用 "new" in find_num.lower() 做子字串比對，會把 "NEW_0" 這種預設流水號也誤判成新料而整批跳過，
        # 導致大量既有料件的 Add/Del 完全沒有反映到 Qty 上。
        if pd.isna(change) or change == "":
            continue
        if find_num.strip().lower() == "new":
            skipped_new_rlc += 1
            continue
        
        add_match = re.search(r"Add\s+([A-Za-z0-9,]+?)(?=,?Del|$)", change)
        if add_match:
            for ref in add_match.group(1).split(","):
                if ref.strip(): result_list.append((pn, "Add", ref.strip()))
        del_match = re.search(r"Del\s+([A-Za-z0-9,]+)", change)
        if del_match:
            for ref in del_match.group(1).split(","):
                if ref.strip(): result_list.append((pn, "Del", ref.strip()))

    if logger:
        logger.info(f"Step3：共產生 {len(result_list)} 筆 Add/Del 明細；{skipped_new_rlc} 筆 RLC 新補料件（Find Num='New'）不重複套用 Add/Del。")

    return pd.DataFrame(result_list, columns=["Item Number", "Action", "Ref Des"]).sort_values(by=["Item Number", "Action"])

def step4_update_qty_and_align(res_df, qty_df, temp_cols, logger=None):
    pn_map = defaultdict(list)
    for _, row in qty_df.iterrows():
        pn_map[str(row.iloc[0]).strip()].append((str(row.iloc[1]).strip(), str(row.iloc[2]).strip()))

    # 在覆寫 Qty 之前，先保留原始 BOM 數量，供輸出「Qty 彙總」分頁使用
    original_qty = res_df["Qty"].copy()

    bad_qty_count = [0]
    def update_qty_func(row):
        pn = str(row.get("Item Number", "")).strip()
        try:
            qty = int(row.get("Qty", 0))
        except Exception:
            qty = 0
            bad_qty_count[0] += 1
        add_count, del_count = 0, 0
        if pn in pn_map:
            for action, _ in pn_map[pn]:
                if action == "Add": qty += 1; add_count += 1
                elif action == "Del": qty -= 1; del_count += 1
        return pd.Series([qty, add_count, del_count])

    res_df[["Qty", "Add_Count", "Del_Count"]] = res_df.apply(update_qty_func, axis=1)
    if logger and bad_qty_count[0] > 0:
        logger.warn(f"有 {bad_qty_count[0]} 筆 Qty 欄位無法轉換為整數（可能是空值或非數字），已預設為 0，請至 Result.xlsx 檢查。")

    # 建立 Qty 彙總表：原始 BOM 數量 -> Add/Del 數量 -> 最終 EDBOM_updated 數量，方便逐筆核對
    qty_summary_df = pd.DataFrame({
        "Find Num": res_df.get("Find Num", ""),
        "Item Number": res_df["Item Number"],
        "原始BOM數量": original_qty,
        "Add數量": res_df["Add_Count"],
        "Del數量": res_df["Del_Count"],
        "最終EDBOM數量": res_df["Qty"],
    })
    if logger:
        logger.info(f"Step4：已產生 Qty 彙總表，共 {len(qty_summary_df)} 筆料件（將輸出至 qty.xlsx 的『Qty彙總』分頁）。")

    missing_cols = [col for col in temp_cols if col not in res_df.columns]
    if logger and missing_cols:
        logger.info(f"套用樣板 (temp.xlsx) 欄位時，補上 {len(missing_cols)} 個空白欄位：{missing_cols}")
    for col in temp_cols:
        if col not in res_df.columns: res_df[col] = "" 
    extra_cols = [c for c in res_df.columns if c not in temp_cols]
    final_df = res_df[temp_cols + extra_cols]
    return final_df, qty_summary_df

def build_change_rich_text(change_str):
    """將『變更差異』欄位的文字轉成 Rich Text：Add 部分藍色、Del 部分紅色，方便肉眼快速辨識新增/刪除。
    沿用 step2/step3 相同的正規表示式來抓取 Add/Del 區段，確保切割位置與判斷邏輯完全一致。"""
    if change_str is None:
        return None
    s = str(change_str).strip()
    if s == "" or s.lower() == "nan":
        return None
    add_match = re.search(r"Add\s+([A-Za-z0-9,]+?)(?=,?Del|$)", s)
    del_match = re.search(r"Del\s+([A-Za-z0-9,]+)", s)
    spans = []
    if add_match: spans.append((add_match.start(), add_match.end(), "0000FF"))  # 藍色
    if del_match: spans.append((del_match.start(), del_match.end(), "FF0000"))  # 紅色
    if not spans:
        return None
    spans.sort(key=lambda x: x[0])
    blocks = []
    cursor = 0
    for start, end, color in spans:
        if start > cursor:
            blocks.append(s[cursor:start])
        blocks.append(TextBlock(InlineFont(color=color), s[start:end]))
        cursor = end
    if cursor < len(s):
        blocks.append(s[cursor:])
    return CellRichText(*blocks)

def apply_template_and_save(final_df, template_path, output_path, logger=None):
    wb = load_workbook(template_path)
    ws = wb.active

    # 重要：必須先把樣板列（第2列）的樣式擷取出來，再執行 delete_rows。
    # 原本的寫法是「先刪除列、再讀取第2列樣式」，但 delete_rows(2, ws.max_row) 會把樣板列本身也一併刪除，
    # 導致之後讀到的都是全新的空白儲存格（has_style 永遠是 False），所以輸出檔案才會完全沒有格線／樣式。
    max_col = ws.max_column
    template_styles = {}
    for j in range(1, max_col + 1):
        src_cell = ws.cell(row=2, column=j)
        if src_cell.has_style:
            template_styles[j] = {
                "font": copy.copy(src_cell.font),
                "border": copy.copy(src_cell.border),
                "fill": copy.copy(src_cell.fill),
                "number_format": src_cell.number_format,
                "alignment": copy.copy(src_cell.alignment),
            }
    if logger:
        logger.info(f"已擷取樣板第 2 列共 {len(template_styles)} 欄的格式（框線/字型/對齊），將套用到所有輸出資料列。")

    ws.delete_rows(2, ws.max_row)
    change_col_idx = None
    if "變更差異" in list(final_df.columns):
        change_col_idx = list(final_df.columns).index("變更差異") + 1  # 1-indexed for openpyxl

    for i, row in final_df.iterrows():
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i + 2, column=j)
            style = template_styles.get(j)
            if j == change_col_idx:
                rich_value = build_change_rich_text(value)
                cell.value = rich_value if rich_value is not None else value
                if style:
                    # 變更差異欄位刻意不覆寫 cell.font，避免蓋掉 Add/Del 的個別上色
                    cell.border, cell.fill, cell.number_format, cell.alignment = style["border"], style["fill"], style["number_format"], style["alignment"]
            else:
                cell.value = value
                if style:
                    cell.font, cell.border, cell.fill, cell.number_format, cell.alignment = style["font"], style["border"], style["fill"], style["number_format"], style["alignment"]
    wb.save(output_path)
    if logger:
        logger.info(f"已套用樣板 '{template_path}'，輸出 {len(final_df)} 筆資料列至 '{output_path}'。")



# =====================================================
# ✅ 主程式執行 (包含 Try-Catch Debug Log 功能)
# =====================================================
if st.button("🚀 開始執行自動化流程", use_container_width=True):
    progress_bar = st.progress(0)
    log_container = st.empty()
    st.session_state.run_complete = False

    maps = {
        "ed_change": ed_map_change, "ed_pn": ed_map_pn, "ed_ref": ed_map_ref,
        "bom_find": bom_map_find, "bom_item": bom_map_item, "bom_desc": bom_map_desc,
        "bom_qty": bom_map_qty, "bom_ref": bom_map_ref, "bom_mfr": bom_map_mfr, "bom_mpn": bom_map_mpn,
        "rlc_item": rlc_map_item, "rlc_desc": rlc_map_desc, "rlc_mfr": rlc_map_mfr, "rlc_mpn": rlc_map_mpn
    }

    log_dir = "log"
    if not os.path.exists(log_dir): os.makedirs(log_dir)
    # 清除舊的 error_trace
    if os.path.exists(os.path.join(log_dir, "error_trace.txt")):
        os.remove(os.path.join(log_dir, "error_trace.txt"))

    logger = RunLogger(log_dir)
    logger.info(f"BOM 來源檔案：{bom_filename}，標題列位置（0-indexed）：{bom_header_row}")
    logger.info(f"欄位對應 - BOM: Find Num={bom_map_find}, Item Number={bom_map_item}, Description={bom_map_desc}, "
                f"Qty={bom_map_qty}, Ref Des={bom_map_ref}, Mfr. Name={bom_map_mfr}, Mfr. Part Number={bom_map_mpn}")
    logger.info(f"欄位對應 - EDlist: Change={ed_map_change}, PN={ed_map_pn}, Ref Des={ed_map_ref}")
    logger.info(f"欄位對應 - RLClist: Item Number={rlc_map_item}, Description={rlc_map_desc}, Mfr. Name={rlc_map_mfr}, Mfr. Part Number={rlc_map_mpn}")

    current_step = "初始化"
    try:
        with st.spinner("自動化流程執行中，請稍候..."):
            # Step 1
            current_step = "Step 1：讀取檔案並比對 EDlist 與 BOM"
            log_container.info("【Step 1】比對 EDlist 與 BOM...")
            ed_df_raw = smart_read_excel_header("EDlist.xlsx", header_row=0, logger=logger)
            bom_df_raw = smart_read_excel_header(bom_filename, header_row=bom_header_row, logger=logger)
            matched_bom, missing_df, ed_df = step1_match_ed_bom(ed_df_raw, bom_df_raw, maps, logger=logger)
            has_newpart = len(missing_df) > 0
            progress_bar.progress(20)

            # Step 1.5 RLC
            current_step = "Step 1.5：RLC 補料"
            df_rlc_found, df_qty = pd.DataFrame(), pd.DataFrame()
            if has_newpart:
                log_container.warning("⚠️ 發現缺料，啟動 RLC 補找流程...")
                export_excel_with_calibri(missing_df, "newpart.xlsx")

                df_rlc_raw = smart_read_excel_header("RLClist.xlsx", header_row=0, logger=logger)
                matched_bom, not_found_df, df_rlc_found = step1_5_rlc_search(missing_df, df_rlc_raw, matched_bom, maps, logger=logger)

                if not not_found_df.empty:
                    export_excel_with_calibri(not_found_df[["Item Number", "Ref Des"]], "RLC_NotFound.xlsx")
            else:
                log_container.success("✅ 無缺料，略過 RLC 流程。")
            progress_bar.progress(40)

            if matched_bom.empty:
                logger.error("最終比對結果 matched_bom 仍為 0 筆，後續步驟產出的 EDBOM_updated.xlsx 將會是空表。請檢查上方欄位對應是否正確。")

            # Step 2
            current_step = "Step 2：計算變更差異並更新 Ref Des"
            log_container.info("【Step 2】計算變更差異並更新 Ref Des...")
            res_df = step2_calc_diff(matched_bom, ed_df, logger=logger)
            logger.info(f"Step 2 完成，res_df 共 {len(res_df)} 筆。")
            progress_bar.progress(60)

            # Step 3
            current_step = "Step 3：計算 Qty 數量差異"
            log_container.info("【Step 3】計算 Qty 數量差異...")
            qty_df = step3_count_qty(res_df, logger=logger)
            logger.info(f"Step 3 完成，偵測到 {len(qty_df)} 筆 Add/Del 明細。")
            progress_bar.progress(80)

            # Step 4
            current_step = "Step 4：更新最終 Qty 並套用樣板格式"
            log_container.info("【Step 4】更新最終 Qty 並套用樣板格式...")
            final_df, qty_summary_df = step4_update_qty_and_align(res_df, qty_df, temp_cols, logger=logger)
            apply_template_and_save(final_df, "temp.xlsx", "EDBOM_updated.xlsx", logger=logger)

            # Step 5
            current_step = "Step 5：寫出系統中繼檔案"
            log_container.info("【Step 5】寫出系統中繼檔案 (Logs)...")
            log_dict = {
                "Result.xlsx": res_df,
                "qty.xlsx": {"AddDel明細": qty_df, "Qty彙總": qty_summary_df},
                "RLC_result.xlsx": df_rlc_found
            }
            save_intermediate_logs(log_dir, log_dict)

            files_to_move = ["EDBOM.xlsx", "qty.xlsx", "Result.xlsx", "result2.xlsx", "EDlistN.xlsx", "RLC_result.xlsx", "newpart.xlsx"]
            for file in files_to_move:
                if os.path.exists(file):
                    dest_path = os.path.join(log_dir, file)
                    if os.path.exists(dest_path): os.remove(dest_path)
                    shutil.move(file, dest_path)

            progress_bar.progress(100)
            logger.info("流程全部執行完畢。")
            logger.save()

            # 打包所有 Log 供下載
            shutil.make_archive("debug_logs", "zip", log_dir)

            st.session_state.run_complete = True
            st.session_state.has_newpart = has_newpart
            st.session_state.show_balloons = True
            st.session_state.debug_log_text = logger.text()

    except Exception as e:
        logger.error(f"執行過程中發生例外，中斷於「{current_step}」：{e}")
        logger.save()
        st.session_state.debug_log_text = logger.text()
        st.error(f"❌ 發生未預期的錯誤（中斷於「{current_step}」）：{str(e)}")

        # 寫入詳細錯誤訊息（含 traceback 與目前為止的詳細日誌）
        with open(os.path.join(log_dir, "error_trace.txt"), "w", encoding="utf-8") as f:
            f.write(f"發生錯誤的步驟：{current_step}\n\n")
            f.write("===== Traceback =====\n")
            f.write(traceback.format_exc())
            f.write("\n\n===== 詳細執行日誌 =====\n")
            f.write(logger.text())

        # 打包 Log 並顯示下載按鈕
        shutil.make_archive("debug_logs", "zip", log_dir)
        with open("debug_logs.zip", "rb") as f:
            st.download_button("📦 下載詳細 Error Log 供工程師除錯", data=f, file_name="debug_logs_error.zip", type="primary")

if st.session_state.debug_log_text:
    with st.expander("🔍 詳細執行日誌 (Debug Log)", expanded=not st.session_state.run_complete):
        st.code(st.session_state.debug_log_text, language="text")

# =====================================================
# ✅ 獨立渲染區塊：完成通知與下載區域
# =====================================================
if st.session_state.run_complete:
    if st.session_state.show_balloons:
        st.success("✅ 所有流程執行完畢！可於下方下載最終檔案，或預覽資料。")
        st.balloons()
        st.session_state.show_balloons = False
    else:
        st.success("✅ 執行結果保留中。如需更新資料，請再次點擊上方按鈕執行。")
    
    # 修改：優化後的缺料提示邏輯
    if os.path.exists("RLC_NotFound.xlsx"):
        st.error("⚠️ **嚴重提醒：有部分料件無法從 RLC 補齊！請務必下載「完全找不到的料件」清單進行人工確認。**")
    elif st.session_state.has_newpart:
        st.info("💡 **提示：初步比對時有缺料，但已全數從 RLC 自動補齊！您仍可下載「缺料原始清單 (newpart)」留存。**")

    st.markdown("### 📥 下載成果")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if os.path.exists("EDBOM_updated.xlsx"):
            with open("EDBOM_updated.xlsx", "rb") as f:
                st.download_button("📥 最終完成表", data=f, file_name="EDBOM_updated.xlsx", use_container_width=True)
    with col2:
        if os.path.exists("RLC_NotFound.xlsx"):
            with open("RLC_NotFound.xlsx", "rb") as f:
                st.download_button("📥 完全找不到的料件", data=f, file_name="RLC_NotFound.xlsx", use_container_width=True)
        elif os.path.exists(os.path.join("log", "newpart.xlsx")) and st.session_state.has_newpart:
            st.info("✅ 缺料已全數從 RLC 補齊！")
    with col3:
        newpart_log = os.path.join("log", "newpart.xlsx")
        if os.path.exists(newpart_log):
            with open(newpart_log, "rb") as f:
                st.download_button("📦 缺料原始清單", data=f, file_name="newpart.xlsx", use_container_width=True)
    with col4:
        if os.path.exists("debug_logs.zip"):
            with open("debug_logs.zip", "rb") as f:
                st.download_button("🛠️ 下載系統 Debug Logs", data=f, file_name="debug_logs.zip", use_container_width=True)
    
    st.markdown("---")
    st.subheader("👁️ 最終結果預覽 (EDBOM_updated)")
    if os.path.exists("EDBOM_updated.xlsx"):
        try:
            st.dataframe(pd.read_excel("EDBOM_updated.xlsx"), use_container_width=True, height=400)
        except Exception:
            st.warning("無法載入預覽資料，請直接下載檔案確認。")
